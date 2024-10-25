from openpyxl import Workbook, load_workbook
import pandas as pd

filename = "Book.xlsx"

try:
  book = load_workbook(filename)
except FileNotFoundError:
  book = Workbook()
  sheet = book.active
  headers = ["Name", "Email", "Phone Number"]
  sheet.append(headers)
  book.save(filename)

sheet = book.active

def add_user():
  name = input("\nName: ")
  email = input("Email: ")
  phone_number = input("Phone Number: ")

  new_data = sheet.max_row + 1

  sheet[f"A{new_data}"] = name
  sheet[f"B{new_data}"] = email
  sheet[f"C{new_data}"] = phone_number

  book.save(filename)

def display_user():
  users = pd.read_excel(filename)
  print(users)

def main():
  while True:
    print("\nMake a choice to proceed:- ")
    print("Press '1' to Add user.")
    print("Press '2' to Display user.")
    print("Press '3' to Exit.")

    choice = input("Press any key to continue: ")

    if choice == '1':
      add_user()
    elif choice == '2':
      display_user()
    elif choice == '3':
      print("\nExiting program.")
      break
    else:
      print("\nInvalid choice, please try again.")

main()